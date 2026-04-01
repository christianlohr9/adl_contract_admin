"""Franchise-tag price validation against the authoritative spreadsheet.

Reads the FT5YO $ tab from the 2026 ADL Contract Admin workbook, extracts
positional averages for EFT, NEFT, and TT tag types, then compares them
against the app's tag price calculations.

Two validation layers:
  1. Positional average validation — app's top-N salary averages vs spreadsheet
  2. Per-player price validation — full tag price (incl 120% floor) and opening bids

Usage:
    uv run python scripts/validate_ft_prices.py [--season 2026] [--output FILE]
"""

from __future__ import annotations

import argparse
import asyncio
import sys
from datetime import datetime, timezone
from decimal import Decimal

from openpyxl import load_workbook
from sqlalchemy import select

# ---------------------------------------------------------------------------
# Ensure src/ is on the path so app imports resolve
# ---------------------------------------------------------------------------
sys.path.insert(0, "src")

from app.core.db import async_session  # noqa: E402
from app.models.contract import Contract  # noqa: E402
from app.models.player import Player  # noqa: E402
from app.services.franchise_tags import (  # noqa: E402
    _get_top_n_positional_salaries,
    calculate_franchise_tags,
    calculate_opening_bid,
    calculate_tag_salary,
    check_tag_eligibility,
)
from app.services.rules import (  # noqa: E402
    ceil_100k,
    floor_100k,
    get_sd_minimum,
    round_to_10k,
)

SPREADSHEET = "2026 ADL Contract Admin.xlsx"
SHEET_NAME = "FT5YO $"

# Positions in column order (cols B-K / N-W)
POSITIONS = ["QB", "RB", "WR", "TE", "PK/PN", "DT", "DE", "LB", "CB", "S"]

# Map spreadsheet position labels to app position codes
POS_MAP = {
    "QB": "QB",
    "RB": "RB",
    "WR": "WR",
    "TE": "TE",
    "PK/PN": "PK",  # App uses PK; PK/PN grouped automatically
    "DT": "DT",
    "DE": "DE",
    "LB": "LB",
    "CB": "CB",
    "S": "S",
}

# Tag types and their row indices in the spreadsheet
# Row 0: header ("February 15 Salaries" / "July 1 Salaries")
# Row 1: position headers
# Row 2: EFT* (Feb) / EFT (Jul)
# Row 3: NEFT
# Row 4: TT
TAG_ROWS = {"EFT": 2, "NEFT": 3, "TT": 4}

# Feb 15 section: cols 1-10 (B-K, 0-indexed)
FEB_COL_START = 1
# July 1 section: cols 13-22 (N-W, 0-indexed)
JUL_COL_START = 13


def load_spreadsheet_prices() -> dict[str, dict[str, dict[str, Decimal | None]]]:
    """Load positional averages from the FT5YO $ tab.

    Returns:
        {section: {tag_type: {position: Decimal}}}
        where section is "feb15" or "jul1"
    """
    wb = load_workbook(SPREADSHEET, read_only=True, data_only=True)
    ws = wb[SHEET_NAME]

    rows_data: list[list] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        rows_data.append(list(row))
        if i > 10:
            break
    wb.close()

    result: dict[str, dict[str, dict[str, Decimal | None]]] = {
        "feb15": {},
        "jul1": {},
    }

    for tag_type, row_idx in TAG_ROWS.items():
        row = rows_data[row_idx]
        result["feb15"][tag_type] = {}
        result["jul1"][tag_type] = {}

        for pos_idx, pos_name in enumerate(POSITIONS):
            # Feb 15 section
            feb_val = row[FEB_COL_START + pos_idx]
            if feb_val is not None:
                result["feb15"][tag_type][pos_name] = Decimal(str(feb_val))
            else:
                result["feb15"][tag_type][pos_name] = None

            # July 1 section
            jul_val = row[JUL_COL_START + pos_idx]
            if jul_val is not None:
                result["jul1"][tag_type][pos_name] = Decimal(str(jul_val))
            else:
                result["jul1"][tag_type][pos_name] = None

    return result


def load_spreadsheet_individual_salaries() -> dict[str, dict[str, list[Decimal]]]:
    """Load individual salary rankings from both sections.

    Returns:
        {section: {position: [salary1, salary2, ...]}}
    """
    wb = load_workbook(SPREADSHEET, read_only=True, data_only=True)
    ws = wb[SHEET_NAME]

    rows_data: list[list] = []
    for row in ws.iter_rows(values_only=True):
        rows_data.append(list(row))
    wb.close()

    result: dict[str, dict[str, list[Decimal]]] = {
        "feb15": {pos: [] for pos in POSITIONS},
        "jul1": {pos: [] for pos in POSITIONS},
    }

    # Individual salaries start at row 7 (0-indexed), rank 1, 2, 3, ...
    for row_idx in range(7, len(rows_data)):
        row = rows_data[row_idx]
        if row[0] is None:
            break

        for pos_idx, pos_name in enumerate(POSITIONS):
            feb_val = row[FEB_COL_START + pos_idx]
            if feb_val is not None:
                result["feb15"][pos_name].append(Decimal(str(feb_val)))

            jul_val = row[JUL_COL_START + pos_idx]
            if jul_val is not None:
                result["jul1"][pos_name].append(Decimal(str(jul_val)))

    return result


async def validate_positional_averages(
    season: int,
) -> tuple[
    list[dict],  # comparison rows
    list[dict],  # discrepancies
    dict[str, dict[str, list[Decimal]]],  # app top-N salaries
]:
    """Layer 1: Compare app positional averages against spreadsheet.

    Returns (comparison_rows, discrepancies, app_salaries).
    """
    ss_prices = load_spreadsheet_prices()
    ss_salaries = load_spreadsheet_individual_salaries()

    comparisons: list[dict] = []
    discrepancies: list[dict] = []
    app_top_salaries: dict[str, dict[str, list[Decimal]]] = {}

    async with async_session() as session:
        for pos_name in POSITIONS:
            app_pos = POS_MAP[pos_name]
            app_top_salaries[pos_name] = {}

            for tag_type in ("EFT", "NEFT", "TT"):
                n = 10 if tag_type == "TT" else 5
                top_sals = await _get_top_n_positional_salaries(
                    session, app_pos, season, n
                )
                app_top_salaries[pos_name][tag_type] = top_sals

                app_avg = (
                    round_to_10k(sum(top_sals) / len(top_sals))
                    if top_sals
                    else Decimal("0")
                )

                # EFT uses Feb 15 in the spreadsheet (marked with *)
                # NEFT/TT use July 1
                # But both sections have all three types; compare against both
                # The app's data should match one section
                feb_val = ss_prices["feb15"][tag_type].get(pos_name)
                jul_val = ss_prices["jul1"][tag_type].get(pos_name)

                # Determine which section matches
                feb_match = feb_val is not None and app_avg == round_to_10k(feb_val)
                jul_match = jul_val is not None and app_avg == round_to_10k(jul_val)

                row = {
                    "position": pos_name,
                    "tag_type": tag_type,
                    "app_avg": app_avg,
                    "feb15_val": feb_val,
                    "jul1_val": jul_val,
                    "feb_match": feb_match,
                    "jul_match": jul_match,
                    "n": n,
                    "count": len(top_sals),
                    "top_sals": top_sals,
                }
                comparisons.append(row)

                # Flag discrepancies — app should match July 1 section
                # (since the DB has current/latest salary data)
                if not jul_match:
                    row_disc = dict(row)
                    # Also check against Feb 15
                    row_disc["matches_feb"] = feb_match
                    discrepancies.append(row_disc)

    return comparisons, discrepancies, app_top_salaries


async def validate_per_player_prices(
    season: int,
) -> tuple[list[dict], list[dict], list[dict]]:
    """Layer 2: Validate per-player tag prices for FT-eligible players.

    Returns (player_results, price_discrepancies, bid_discrepancies).
    """
    sd_minimum = get_sd_minimum(season)
    player_results: list[dict] = []
    price_discrepancies: list[dict] = []
    bid_discrepancies: list[dict] = []

    async with async_session() as session:
        # Find all players with expired contracts (FT-eligible candidates)
        result = await session.execute(
            select(Contract.player_id, Contract.team_id)
            .where(
                Contract.season == season,
                Contract.years_remaining == 0,
            )
            .distinct()
        )
        candidates = result.all()
        print(
            f"Found {len(candidates)} players with expired contracts",
            file=sys.stderr,
        )

        for player_id, team_id in candidates:
            eligible, reason = await check_tag_eligibility(
                session, player_id, season, team_id
            )
            if not eligible:
                continue

            ft_result = await calculate_franchise_tags(
                session, player_id, season, team_id
            )
            if not ft_result.eligible:
                continue

            prev_salary = ft_result.previous_salary
            salary_floor = Decimal("1.20") * prev_salary

            player_info = {
                "player_id": player_id,
                "player_name": ft_result.player_name,
                "position": ft_result.position,
                "prev_salary": prev_salary,
                "salary_floor": round_to_10k(salary_floor),
                "options": [],
            }

            for opt in ft_result.options:
                # Recalculate independently to verify
                app_pos = ft_result.position
                n = 10 if opt.tag_type == "TT" else 5
                top_sals = await _get_top_n_positional_salaries(
                    session, app_pos, season, n
                )
                pos_avg = (
                    sum(top_sals) / len(top_sals) if top_sals else Decimal("0")
                )
                expected_salary = round_to_10k(max(pos_avg, salary_floor))
                expected_bid = None
                if opt.tag_type in ("NEFT", "TT"):
                    expected_bid = calculate_opening_bid(opt.salary, sd_minimum)

                opt_info = {
                    "tag_type": opt.tag_type,
                    "salary": opt.salary,
                    "expected_salary": expected_salary,
                    "pos_avg": round_to_10k(pos_avg),
                    "salary_floor_used": round_to_10k(salary_floor) > round_to_10k(pos_avg),
                    "opening_bid": opt.opening_bid,
                    "expected_bid": expected_bid,
                }
                player_info["options"].append(opt_info)

                if opt.salary != expected_salary:
                    price_discrepancies.append(
                        {
                            "player_name": ft_result.player_name,
                            "position": ft_result.position,
                            "tag_type": opt.tag_type,
                            "app_salary": opt.salary,
                            "expected_salary": expected_salary,
                            "pos_avg": round_to_10k(pos_avg),
                            "salary_floor": round_to_10k(salary_floor),
                            "prev_salary": prev_salary,
                        }
                    )

                if opt.opening_bid is not None and expected_bid is not None:
                    if opt.opening_bid != expected_bid:
                        bid_discrepancies.append(
                            {
                                "player_name": ft_result.player_name,
                                "tag_type": opt.tag_type,
                                "app_bid": opt.opening_bid,
                                "expected_bid": expected_bid,
                                "tag_salary": opt.salary,
                                "sd_minimum": sd_minimum,
                            }
                        )

            player_results.append(player_info)

    return player_results, price_discrepancies, bid_discrepancies


def generate_report(
    season: int,
    comparisons: list[dict],
    pos_discrepancies: list[dict],
    player_results: list[dict],
    price_discrepancies: list[dict],
    bid_discrepancies: list[dict],
    sd_minimum: Decimal,
) -> str:
    """Generate a markdown validation report."""
    now = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    lines: list[str] = []

    lines.append("# Phase 18-02: Franchise Tag Price Validation Report")
    lines.append("")
    lines.append(f"**Validated:** {now}")
    lines.append(f"**Method:** Spreadsheet vs app comparison via `scripts/validate_ft_prices.py`")
    lines.append(f"**Source:** `{SPREADSHEET}` -- `{SHEET_NAME}` tab")
    lines.append(f"**Season:** {season}")
    lines.append(f"**SD Minimum:** {sd_minimum}")
    lines.append("")
    lines.append("---")
    lines.append("")

    # -----------------------------------------------------------------------
    # Layer 1: Positional Averages
    # -----------------------------------------------------------------------
    lines.append("## Layer 1: Positional Average Validation")
    lines.append("")
    lines.append(
        "Compares the app's computed top-N salary averages against the spreadsheet's "
        "EFT/NEFT/TT rows. The spreadsheet has two sections: February 15 salaries "
        "(used for EFT*) and July 1 salaries (used for NEFT/TT). The app uses "
        "season-1 contract data from the database."
    )
    lines.append("")

    # Summary table
    lines.append("### Positional Average Comparison")
    lines.append("")
    lines.append("| Position | Tag | Top-N | App Avg | SS Feb15 | SS Jul1 | Feb Match | Jul Match |")
    lines.append("|----------|-----|-------|---------|----------|---------|-----------|-----------|")
    for c in comparisons:
        feb_str = str(c["feb15_val"]) if c["feb15_val"] is not None else "-"
        jul_str = str(c["jul1_val"]) if c["jul1_val"] is not None else "-"
        feb_m = "Y" if c["feb_match"] else "N"
        jul_m = "Y" if c["jul_match"] else "N"
        lines.append(
            f"| {c['position']} | {c['tag_type']} | {c['count']}/{c['n']} "
            f"| {c['app_avg']} | {feb_str} | {jul_str} | {feb_m} | {jul_m} |"
        )
    lines.append("")

    # Discrepancies
    total_comparisons = len(comparisons)
    disc_count = len(pos_discrepancies)
    match_count = total_comparisons - disc_count
    lines.append(f"**Results:** {match_count}/{total_comparisons} positional averages match July 1 section")
    lines.append("")

    if pos_discrepancies:
        lines.append("### Positional Average Discrepancies (vs July 1)")
        lines.append("")
        for d in pos_discrepancies:
            lines.append(f"#### {d['position']} - {d['tag_type']}")
            lines.append(f"- **App average:** {d['app_avg']}")
            lines.append(f"- **SS Jul1:** {d['jul1_val']}")
            lines.append(f"- **SS Feb15:** {d['feb15_val']}")
            lines.append(f"- **Matches Feb15:** {'Yes' if d.get('matches_feb') else 'No'}")
            lines.append(f"- **Top {d['n']} salaries ({d['count']} found):** {[str(s) for s in d['top_sals']]}")
            lines.append("")
    else:
        lines.append("No positional average discrepancies found against July 1 section.")
        lines.append("")

    lines.append("---")
    lines.append("")

    # -----------------------------------------------------------------------
    # Layer 2: Per-Player Prices
    # -----------------------------------------------------------------------
    lines.append("## Layer 2: Per-Player Price Validation")
    lines.append("")
    lines.append(
        f"Validated {len(player_results)} FT-eligible players. For each player, "
        f"verifies tag salary = MAX(positional_avg, 1.20 x prev_salary) and "
        f"opening bid = MAX(CEIL_100K(SD_Min - 0.1), FLOOR_100K(tag_salary))."
    )
    lines.append("")

    # Summary
    lines.append(f"**Price discrepancies:** {len(price_discrepancies)}")
    lines.append(f"**Opening bid discrepancies:** {len(bid_discrepancies)}")
    lines.append("")

    # Player table
    if player_results:
        lines.append("### Per-Player Tag Prices")
        lines.append("")
        lines.append("| Player | Pos | Prev$ | EFT | NEFT | NEFT Bid | TT | TT Bid | Floor Used? |")
        lines.append("|--------|-----|-------|-----|------|----------|----|--------|-------------|")
        for p in sorted(player_results, key=lambda x: x["player_name"]):
            eft = neft = tt = neft_bid = tt_bid = "-"
            floor_used = []
            for opt in p["options"]:
                if opt["tag_type"] == "EFT":
                    eft = str(opt["salary"])
                    if opt["salary_floor_used"]:
                        floor_used.append("EFT")
                elif opt["tag_type"] == "NEFT":
                    neft = str(opt["salary"])
                    neft_bid = str(opt["opening_bid"]) if opt["opening_bid"] else "-"
                    if opt["salary_floor_used"]:
                        floor_used.append("NEFT")
                elif opt["tag_type"] == "TT":
                    tt = str(opt["salary"])
                    tt_bid = str(opt["opening_bid"]) if opt["opening_bid"] else "-"
                    if opt["salary_floor_used"]:
                        floor_used.append("TT")
            floor_str = ", ".join(floor_used) if floor_used else "-"
            lines.append(
                f"| {p['player_name']} | {p['position']} | {p['prev_salary']} "
                f"| {eft} | {neft} | {neft_bid} | {tt} | {tt_bid} | {floor_str} |"
            )
        lines.append("")

    # Price discrepancies
    if price_discrepancies:
        lines.append("### Price Discrepancies")
        lines.append("")
        lines.append("| Player | Pos | Tag | App$ | Expected$ | PosAvg | 120%Floor | Prev$ |")
        lines.append("|--------|-----|-----|------|-----------|--------|-----------|-------|")
        for d in price_discrepancies:
            lines.append(
                f"| {d['player_name']} | {d['position']} | {d['tag_type']} "
                f"| {d['app_salary']} | {d['expected_salary']} "
                f"| {d['pos_avg']} | {d['salary_floor']} | {d['prev_salary']} |"
            )
        lines.append("")

    # Bid discrepancies
    if bid_discrepancies:
        lines.append("### Opening Bid Discrepancies")
        lines.append("")
        lines.append("| Player | Tag | App Bid | Expected Bid | Tag$ | SD Min |")
        lines.append("|--------|-----|---------|--------------|------|--------|")
        for d in bid_discrepancies:
            lines.append(
                f"| {d['player_name']} | {d['tag_type']} "
                f"| {d['app_bid']} | {d['expected_bid']} "
                f"| {d['tag_salary']} | {d['sd_minimum']} |"
            )
        lines.append("")

    lines.append("---")
    lines.append("")

    # -----------------------------------------------------------------------
    # Opening Bid Verification
    # -----------------------------------------------------------------------
    lines.append("## Opening Bid Verification")
    lines.append("")
    lines.append(f"SD Minimum for {season}: {sd_minimum}")
    lines.append(f"CEIL_100K(SD_Min - 0.1) = CEIL_100K({sd_minimum} - 0.1) = {ceil_100k(sd_minimum - Decimal('0.1'))}")
    lines.append("")

    bid_checks = []
    for p in player_results:
        for opt in p["options"]:
            if opt["opening_bid"] is not None:
                sd_component = ceil_100k(sd_minimum - Decimal("0.1"))
                tag_component = floor_100k(opt["salary"])
                expected = max(sd_component, tag_component)
                bid_checks.append(
                    {
                        "player": p["player_name"],
                        "tag": opt["tag_type"],
                        "salary": opt["salary"],
                        "sd_comp": sd_component,
                        "tag_comp": tag_component,
                        "expected": expected,
                        "actual": opt["opening_bid"],
                        "match": opt["opening_bid"] == expected,
                    }
                )

    if bid_checks:
        lines.append("| Player | Tag | Salary | SD Comp | Tag Comp | Expected Bid | Actual Bid | Match |")
        lines.append("|--------|-----|--------|---------|----------|--------------|------------|-------|")
        for b in bid_checks:
            m = "Y" if b["match"] else "**N**"
            lines.append(
                f"| {b['player']} | {b['tag']} | {b['salary']} "
                f"| {b['sd_comp']} | {b['tag_comp']} | {b['expected']} "
                f"| {b['actual']} | {m} |"
            )
        lines.append("")

    lines.append("---")
    lines.append("")

    # -----------------------------------------------------------------------
    # Confidence Statement
    # -----------------------------------------------------------------------
    lines.append("## Confidence Statement")
    lines.append("")

    total_issues = len(pos_discrepancies) + len(price_discrepancies) + len(bid_discrepancies)
    if total_issues == 0:
        lines.append(
            f"All {match_count} positional averages match the spreadsheet's July 1 section. "
            f"All {len(player_results)} FT-eligible players have correct tag prices and "
            f"opening bids. The app's franchise tag price calculations are fully consistent "
            f"with the commissioner's spreadsheet."
        )
    else:
        lines.append(
            f"Validation found {len(pos_discrepancies)} positional average discrepancies, "
            f"{len(price_discrepancies)} per-player price discrepancies, and "
            f"{len(bid_discrepancies)} opening bid discrepancies. "
            f"Review the tables above for details."
        )
    lines.append("")
    lines.append("---")
    lines.append("")
    lines.append("*Phase: 18-franchise-tags*")
    lines.append(f"*Validated: {now}*")
    lines.append("")

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Validate FT tag prices: spreadsheet vs app.",
    )
    parser.add_argument(
        "--season",
        type=int,
        default=2026,
        help="ADL season year to validate (default: 2026)",
    )
    parser.add_argument(
        "--output",
        type=str,
        default=None,
        help="Write report to this file (default: stdout)",
    )
    return parser.parse_args()


async def main() -> None:
    args = parse_args()
    sd_minimum = get_sd_minimum(args.season)

    print(f"Starting FT price validation for season {args.season}...", file=sys.stderr)
    print(f"SD Minimum: {sd_minimum}", file=sys.stderr)

    # Layer 1
    print("Layer 1: Validating positional averages...", file=sys.stderr)
    comparisons, pos_discrepancies, app_salaries = await validate_positional_averages(
        args.season
    )
    print(
        f"  {len(comparisons)} comparisons, {len(pos_discrepancies)} discrepancies",
        file=sys.stderr,
    )

    # Layer 2
    print("Layer 2: Validating per-player prices...", file=sys.stderr)
    player_results, price_discrepancies, bid_discrepancies = (
        await validate_per_player_prices(args.season)
    )
    print(
        f"  {len(player_results)} eligible players, "
        f"{len(price_discrepancies)} price discrepancies, "
        f"{len(bid_discrepancies)} bid discrepancies",
        file=sys.stderr,
    )

    report = generate_report(
        args.season,
        comparisons,
        pos_discrepancies,
        player_results,
        price_discrepancies,
        bid_discrepancies,
        sd_minimum,
    )

    if args.output:
        import os

        os.makedirs(os.path.dirname(args.output) or ".", exist_ok=True)
        with open(args.output, "w") as f:
            f.write(report)
        print(f"Report written to {args.output}", file=sys.stderr)
    else:
        print(report)

    total_issues = (
        len(pos_discrepancies) + len(price_discrepancies) + len(bid_discrepancies)
    )
    print(
        f"Done: {total_issues} total discrepancies "
        f"({len(pos_discrepancies)} positional, {len(price_discrepancies)} price, "
        f"{len(bid_discrepancies)} bid)",
        file=sys.stderr,
    )


if __name__ == "__main__":
    asyncio.run(main())
